#!/usr/bin/env python3
"""Import verified player historical season totals from an approved sheet export.

This command intentionally imports a checked/exported CSV or XLSX file, not a
live Google Sheet.  The source workbook stays human-editable while each import
has a reproducible file hash and a row-level reconciliation report.

Only exact player identity keys (name + current team + fantasy position) are
accepted by default.  A short, reviewable alias table covers verified spelling,
suffix, and word-order differences in the source workbook; no fuzzy matching or
estimated statistics are used.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from collections import defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from collegefootballfantasy_api.app.db.model_registry import ensure_models_registered
from collegefootballfantasy_api.app.db.session import SessionLocal
from collegefootballfantasy_api.app.models.college_team import CollegeTeam
from collegefootballfantasy_api.app.models.historical_stats import (
    HistoricalStatImportRun,
    PlayerHistoricalSeasonStat,
)
from collegefootballfantasy_api.app.models.player import Player
from collegefootballfantasy_api.app.models.provider_identity import PlayerProviderId, ProviderIdentityAudit
from collegefootballfantasy_api.app.domain.scoring_engine import calculate_player_fantasy_points
from collegefootballfantasy_api.app.domain.scoring_rules import BETA_KICKER_RULES
from collegefootballfantasy_api.app.services.historical_stats import canonical_json_hash
from collegefootballfantasy_api.app.services.power4 import resolve_power4_school


SOURCE_PROVIDER = "google_season_stats"
SOURCE_TYPE = "google_sheet_export"
PARSER_VERSION = "google-season-stats-v1"
SEASON_TYPE = "regular"
REQUIRED_HEADERS = {"CURRENT TEAM", "DEPTH POS", "PLAYER", "SEASON", "COLLEGE TEAM"}


def _normalized(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return "".join(character for character in text.lower() if character.isalnum())


def _text(value: object) -> str | None:
    text = str(value or "").strip()
    return text or None


def _position(value: object) -> str | None:
    # Depth slots are commonly QB1, RB2, WRX, or TE1.  A word boundary does
    # not exist between a letter and a digit, so accept either one of those
    # suffixes or a true boundary after the fantasy position.
    match = re.match(r"\s*(QB|RB|WR|TE|K)(?=\d|[A-Z]|\b)", str(value or ""), flags=re.IGNORECASE)
    return match.group(1).upper() if match else None


def _number(value: object) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _integer(value: object) -> int | None:
    number = _number(value)
    if number is None or not number.is_integer():
        return None
    return int(number)


def _identity_key(name: str | None, team: str | None, position: str | None) -> tuple[str, str, str]:
    # Sources legitimately use abbreviations such as CAL and WVU while the
    # canonical registry stores California and West Virginia.  Normalize the
    # team component through the approved Power 4 mapping before comparing
    # identities; do not loosen the name or position match.
    canonical_team = resolve_power4_school(team or "") or team
    return (_normalized(name), _normalized(canonical_team), (position or "").upper())


# These source spellings were verified against the app's current canonical
# player records by full source team and fantasy position.  Do not add a name
# here unless it is a clear typo/suffix/nickname/word-order variant of the same
# player.  Ambiguous names must remain in the generated review report.
VERIFIED_SOURCE_NAME_ALIASES: dict[tuple[str, str, str], str] = {
    _identity_key("Aidan Mizell", "UCLA", "WR"): "Aiden Mizell",
    _identity_key("Bryan Jackson II", "Wisconsin", "RB"): "Bryan Jackson",
    _identity_key("Cameron Ball", "West Virginia", "TE"): "Cam Ball",
    _identity_key("Cameron Kossmann", "Boston College", "TE"): "Cameron Kossman",
    _identity_key("Harry Dalton III", "Maryland", "RB"): "Harry Dalton",
    _identity_key("Jaime Ffrench Jr.", "Michigan", "WR"): "Jaime Ffrench",
    _identity_key("Joshua Phifer", "UCLA", "TE"): "Josh Phifer",
    _identity_key("Karle Lacey Jr.", "Texas", "QB"): 'Karle "KJ" Lacey Jr.',
    _identity_key("Na'eem Abdul-Rahim Gladding", "Maryland", "WR"): "Naeem Gladding Abdul-Rahim",
    _identity_key("Traville Frederick Jr.", "Houston", "TE"): "Traville Fredrick Jr.",
}


@dataclass(frozen=True)
class SourceSeasonRow:
    row_number: int
    current_team: str
    depth_position: str
    player_name: str
    season: int | None
    college_team: str | None
    passing_completions: float | None
    passing_attempts: float | None
    passing_yards: float | None
    passing_touchdowns: float | None
    interceptions: float | None
    rushing_attempts: float | None
    rushing_yards: float | None
    rushing_touchdowns: float | None
    receptions: float | None
    receiving_yards: float | None
    receiving_touchdowns: float | None
    field_goals_made: float | None
    field_goals_attempted: float | None
    extra_points_made: float | None
    extra_points_attempted: float | None
    kick_points: float | None
    games_played: int | None = None
    espn_player_id: str | None = None
    source_url: str | None = None

    @property
    def position(self) -> str | None:
        return _position(self.depth_position)

    @property
    def identity_key(self) -> tuple[str, str, str]:
        return _identity_key(self.player_name, self.current_team, self.position)

    def source_payload(self) -> dict[str, object]:
        return asdict(self)


def _headers_from_row(values: Iterable[object]) -> list[str]:
    return [str(value or "").strip().upper() for value in values]


def _row_from_mapping(row_number: int, values: dict[str, object]) -> SourceSeasonRow:
    value = lambda header: values.get(header, "")
    return SourceSeasonRow(
        row_number=row_number,
        current_team=_text(value("CURRENT TEAM")) or "",
        depth_position=_text(value("DEPTH POS")) or "",
        player_name=_text(value("PLAYER")) or "",
        season=_integer(value("SEASON")),
        college_team=_text(value("COLLEGE TEAM")),
        passing_completions=_number(value("PASS CMP")),
        passing_attempts=_number(value("PASS ATT")),
        passing_yards=_number(value("PASS YDS")),
        passing_touchdowns=_number(value("PASS TD")),
        interceptions=_number(value("INT")),
        rushing_attempts=_number(value("RUSH CAR")),
        rushing_yards=_number(value("RUSH YDS")),
        rushing_touchdowns=_number(value("RUSH TD")),
        receptions=_number(value("REC")),
        receiving_yards=_number(value("REC YDS")),
        receiving_touchdowns=_number(value("REC TD")),
        field_goals_made=_number(value("FGM")),
        field_goals_attempted=_number(value("FGA")),
        extra_points_made=_number(value("XPM")),
        extra_points_attempted=_number(value("XPA")),
        kick_points=_number(value("KICK PTS")),
        games_played=_integer(value("GAMES PLAYED") or value("GP")),
        espn_player_id=_text(value("ESPN ID") or value("ESPN PLAYER ID")),
        source_url=_text(value("SOURCE URL")),
    )


def _read_csv(path: Path) -> list[SourceSeasonRow]:
    with path.open(newline="", encoding="utf-8-sig") as source:
        rows = list(csv.reader(source))
    return _rows_from_matrix(rows)


def _read_xlsx(path: Path) -> list[SourceSeasonRow]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - protected by project dependency
        raise RuntimeError("XLSX imports require the openpyxl dependency.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            matrix = list(worksheet.iter_rows(values_only=True))
            try:
                return _rows_from_matrix(matrix)
            except ValueError:
                continue
    finally:
        workbook.close()
    raise ValueError(f"No worksheet in {path.name} contains the required Season Stats headers.")


def _rows_from_matrix(matrix: list[Iterable[object]]) -> list[SourceSeasonRow]:
    header_index = next(
        (
            index
            for index, values in enumerate(matrix[:25])
            if REQUIRED_HEADERS.issubset(set(_headers_from_row(values)))
        ),
        None,
    )
    if header_index is None:
        raise ValueError("Could not find required Season Stats headers in the first 25 rows.")
    headers = _headers_from_row(matrix[header_index])
    parsed: list[SourceSeasonRow] = []
    for source_index, values in enumerate(matrix[header_index + 1 :], start=header_index + 2):
        mapping = {header: values[index] if index < len(values) else "" for index, header in enumerate(headers)}
        if not any(_text(cell) for cell in mapping.values()):
            continue
        parsed.append(_row_from_mapping(source_index, mapping))
    return parsed


def read_source_rows(path: Path) -> list[SourceSeasonRow]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    raise ValueError("Supported source exports are .csv and .xlsx.")


def _file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _team_lookup(teams: Iterable[CollegeTeam]) -> dict[str, CollegeTeam]:
    return {_normalized(team.name): team for team in teams}


def _current_source_key(row: SourceSeasonRow) -> tuple[str, str, str]:
    return row.identity_key


def _resolve_player(
    row: SourceSeasonRow,
    players_by_key: dict[tuple[str, str, str], Player],
) -> tuple[Player | None, str]:
    direct = players_by_key.get(_current_source_key(row))
    if direct:
        return direct, "exact"

    alias_name = VERIFIED_SOURCE_NAME_ALIASES.get(_current_source_key(row))
    if alias_name:
        alias_key = _identity_key(alias_name, row.current_team, row.position)
        resolved = players_by_key.get(alias_key)
        if resolved:
            return resolved, "verified_alias"
    return None, "unmatched"


def _source_provider_id(row: SourceSeasonRow) -> str | None:
    """Only accept an explicit, source-verified ESPN identifier."""
    if row.espn_player_id and row.espn_player_id.isdigit():
        return row.espn_player_id
    # The reviewed workbook also stores ESPN athlete URLs for some legacy
    # rows. The ID segment is still an explicit workbook value; no provider is
    # contacted and arbitrary non-ESPN URLs are never interpreted as IDs.
    if row.source_url and re.match(r"https?://(?:www\.)?espn\.com/", row.source_url, re.IGNORECASE):
        match = re.search(r"/(?:id/)?(\d+)(?:[/?#-]|$)", row.source_url)
        if match:
            return match.group(1)
    return None


def _canonical_fantasy_points(source: SourceSeasonRow) -> tuple[float | None, str | None]:
    """Calculate only from source fields proven by the beta scoring policy."""

    if source.position == "K":
        if source.field_goals_made is None or source.extra_points_made is None:
            return None, None
        points, _ = calculate_player_fantasy_points(
            {"fg_made_0_30": source.field_goals_made, "xp_made": source.extra_points_made},
            BETA_KICKER_RULES,
            source.position,
        )
        return points, "component_stats_canonical_scoring_v2_beta_flat_kicker"
    if all(value is not None for value in (
        source.passing_yards, source.passing_touchdowns, source.interceptions,
        source.rushing_yards, source.rushing_touchdowns, source.receptions,
        source.receiving_yards, source.receiving_touchdowns,
    )):
        points, _ = calculate_player_fantasy_points(
            {
                "pass_yards": source.passing_yards,
                "pass_tds": source.passing_touchdowns,
                "interceptions": source.interceptions,
                "rush_yards": source.rushing_yards,
                "rush_tds": source.rushing_touchdowns,
                "receptions": source.receptions,
                "rec_yards": source.receiving_yards,
                "rec_tds": source.receiving_touchdowns,
            },
            {},
            source.position,
        )
        return points, "component_stats_canonical_scoring_v2_beta_flat_kicker"
    return None, None


def _assign_row(
    target: PlayerHistoricalSeasonStat,
    source: SourceSeasonRow,
    player: Player,
    teams_by_normalized_name: dict[str, CollegeTeam],
    source_hash: str,
    imported_at: datetime,
) -> None:
    historical_canonical_name = resolve_power4_school(source.college_team or "")
    historical_team = teams_by_normalized_name.get(_normalized(historical_canonical_name)) if historical_canonical_name else None
    current_canonical_name = resolve_power4_school(player.school)
    current_team = teams_by_normalized_name.get(_normalized(current_canonical_name)) if current_canonical_name else None
    provider_id = _source_provider_id(source)
    target.provider_player_id = provider_id or f"source-row:{source.row_number}"
    target.team_id = historical_team.id if historical_team else None
    target.historical_team_id = historical_team.id if historical_team else None
    target.current_team_at_import_id = current_team.id if current_team else None
    target.team_name = source.college_team
    target.position = source.position
    target.source_depth_position = source.depth_position
    target.canonical_position = player.position
    target.passing_completions = source.passing_completions
    target.passing_attempts = source.passing_attempts
    target.passing_yards = source.passing_yards
    target.passing_touchdowns = source.passing_touchdowns
    target.interceptions = source.interceptions
    target.rushing_attempts = source.rushing_attempts
    target.rushing_yards = source.rushing_yards
    target.rushing_touchdowns = source.rushing_touchdowns
    target.receptions = source.receptions
    target.receiving_yards = source.receiving_yards
    target.receiving_touchdowns = source.receiving_touchdowns
    target.field_goals_made = source.field_goals_made
    target.field_goals_attempted = source.field_goals_attempted
    target.extra_points_made = source.extra_points_made
    target.extra_points_attempted = source.extra_points_attempted
    target.kick_points = source.kick_points
    target.fantasy_points, target.scoring_rules_version = _canonical_fantasy_points(source)
    target.games_played = source.games_played
    target.fantasy_points_per_game = (
        target.fantasy_points / source.games_played
        if target.fantasy_points is not None and source.games_played and source.games_played > 0
        else None
    )
    target.source_response_hash = canonical_json_hash(source.source_payload())
    target.source_url = source.source_url
    target.source_external_player_id = provider_id
    target.source_type = SOURCE_TYPE
    target.source_modified_at = None
    target.import_version = f"{PARSER_VERSION}:{source_hash[:12]}"
    target.parser_version = PARSER_VERSION
    target.imported_at = imported_at
    target.provider_updated_at = None
    target.raw_labels = {
        "current_team": source.current_team,
        "depth_position": source.depth_position,
        "college_team": source.college_team,
        "source_row": source.row_number,
    }
    target.unknown_labels = {"fantasy_points": "missing_canonical_component"} if target.fantasy_points is None else None
    target.is_final = True


def build_report(source_rows: list[SourceSeasonRow], players: Iterable[Player]) -> dict[str, Any]:
    players_by_key = {_identity_key(player.name, player.school, player.position): player for player in players}
    season_rows = [row for row in source_rows if row.season and row.season >= 1900]
    source_identities_without_season = {
        row.identity_key for row in source_rows if not row.season or row.season < 1900
    }
    matched_source_keys: set[tuple[str, str, str]] = set()
    exact_matches = alias_matches = 0
    unmatched_rows: list[dict[str, object]] = []
    provider_ids_by_player: dict[tuple[str, str, str], set[str]] = defaultdict(set)
    players_by_provider_id: dict[str, set[tuple[str, str, str]]] = defaultdict(set)
    for row in season_rows:
        player, match_type = _resolve_player(row, players_by_key)
        if player is None:
            unmatched_rows.append(
                {
                    "row_number": row.row_number,
                    "player_name": row.player_name,
                    "current_team": row.current_team,
                    "position": row.position,
                    "season": row.season,
                    "reason": "no exact canonical player identity match",
                }
            )
            continue
        matched_source_keys.add(_identity_key(player.name, player.school, player.position))
        exact_matches += int(match_type == "exact")
        alias_matches += int(match_type == "verified_alias")
        provider_id = _source_provider_id(row)
        if provider_id:
            canonical_key = _identity_key(player.name, player.school, player.position)
            provider_ids_by_player[canonical_key].add(provider_id)
            players_by_provider_id[provider_id].add(canonical_key)
    catalog_without_history = []
    for player in players:
        player_key = _identity_key(player.name, player.school, player.position)
        if player_key in matched_source_keys:
            continue
        # The source marks incoming freshmen and other players without a
        # completed college season using an empty/zero season.  Keep that
        # distinct from a missing identity row so the UI can state the truth
        # without implying an import failure.
        reason = (
            "source_has_no_completed_college_season"
            if player_key in source_identities_without_season
            else "no_source_identity_row"
        )
        catalog_without_history.append(
            {
                "player_id": player.id,
                "player_name": player.name,
                "team": player.school,
                "position": player.position,
                "reason": reason,
            }
        )
    provider_id_conflicts = [
        {
            "provider": "espn",
            "provider_player_id": provider_id,
            "canonical_player_keys": ["|".join(key) for key in sorted(player_keys)],
            "reason": "one_trusted_espn_id_maps_to_multiple_canonical_players",
        }
        for provider_id, player_keys in sorted(players_by_provider_id.items())
        if len(player_keys) > 1
    ]
    player_provider_id_conflicts = [
        {
            "canonical_player_key": "|".join(player_key),
            "provider": "espn",
            "provider_player_ids": sorted(provider_ids),
            "reason": "one_canonical_player_has_multiple_trusted_espn_ids",
        }
        for player_key, provider_ids in sorted(provider_ids_by_player.items())
        if len(provider_ids) > 1
    ]
    return {
        "source_rows": len(source_rows),
        "season_rows": len(season_rows),
        "non_season_or_blank_rows": len(source_rows) - len(season_rows),
        "exact_source_rows_matched": exact_matches,
        "verified_alias_source_rows_matched": alias_matches,
        "unmatched_source_rows": unmatched_rows,
        "catalog_players_without_source_history": catalog_without_history,
        "catalog_players_without_source_history_by_reason": {
            "source_has_no_completed_college_season": sum(
                item["reason"] == "source_has_no_completed_college_season" for item in catalog_without_history
            ),
            "no_source_identity_row": sum(
                item["reason"] == "no_source_identity_row" for item in catalog_without_history
            ),
        },
        "trusted_espn_id_player_count": len(provider_ids_by_player),
        "trusted_espn_id_conflicts": provider_id_conflicts,
        "trusted_espn_id_player_conflicts": player_provider_id_conflicts,
        "trusted_espn_id_conflict_count": len(provider_id_conflicts) + len(player_provider_id_conflicts),
    }


def import_rows(path: Path, *, apply: bool) -> dict[str, Any]:
    source_hash = _file_hash(path)
    source_rows = read_source_rows(path)
    ensure_models_registered()
    with SessionLocal() as db:
        players = db.query(Player).order_by(Player.id).all()
        report = build_report(source_rows, players)
        report.update({"source_path": str(path), "source_sha256": source_hash, "apply": apply})
        if not apply:
            return report
        if report["trusted_espn_id_conflict_count"]:
            raise ValueError(
                "Trusted ESPN identity reconciliation is blocked by "
                f"{report['trusted_espn_id_conflict_count']} source conflict(s)."
            )

        players_by_key = {_identity_key(player.name, player.school, player.position): player for player in players}
        teams_by_normalized_name = _team_lookup(db.query(CollegeTeam).all())
        rows_to_import = [row for row in source_rows if row.season and row.season >= 1900]
        run = HistoricalStatImportRun(
            provider=SOURCE_PROVIDER,
            requested_seasons=sorted({row.season for row in rows_to_import if row.season}),
            requested_player_ids=[],
            status="running",
            started_at=datetime.now(timezone.utc),
            players_requested=0,
            trigger_type="manual",
        )
        db.add(run)
        db.flush()
        imported_at = datetime.now(timezone.utc)
        touched_players: set[int] = set()
        rows_inserted = rows_updated = 0
        mappings_inserted = mappings_unchanged = 0
        errors: list[dict[str, object]] = []
        for source in rows_to_import:
            player, match_type = _resolve_player(source, players_by_key)
            if not player:
                continue
            provider_id = _source_provider_id(source)
            if source.espn_player_id and not provider_id:
                raise ValueError(f"Malformed verified ESPN ID in source row {source.row_number}")
            if provider_id:
                by_provider_id = db.query(PlayerProviderId).filter_by(provider="espn", provider_player_id=provider_id).one_or_none()
                by_player = db.query(PlayerProviderId).filter_by(provider="espn", player_id=player.id).one_or_none()
                if (by_provider_id and by_provider_id.player_id != player.id) or (by_player and by_player.provider_player_id != provider_id):
                    raise ValueError(f"ESPN identity conflict for source row {source.row_number}")
                if by_provider_id is None:
                    db.add(PlayerProviderId(player_id=player.id, provider="espn", provider_player_id=provider_id, match_confidence=1.0, verification_status="verified", verified_at=imported_at))
                    db.add(ProviderIdentityAudit(entity_type="player", entity_id=player.id, action="source_workbook_import", provider="espn", provider_player_id=provider_id, after_state={"source_hash": source_hash, "source_row": source.row_number}, reason="approved previous-stats workbook import"))
                    mappings_inserted += 1
                else:
                    mappings_unchanged += 1
            existing = (
                db.query(PlayerHistoricalSeasonStat)
                .filter(
                    PlayerHistoricalSeasonStat.player_id == player.id,
                    PlayerHistoricalSeasonStat.provider == SOURCE_PROVIDER,
                    PlayerHistoricalSeasonStat.season == source.season,
                    PlayerHistoricalSeasonStat.season_type == SEASON_TYPE,
                    PlayerHistoricalSeasonStat.team_name == source.college_team,
                )
                .one_or_none()
            )
            if existing is None:
                existing = PlayerHistoricalSeasonStat(
                    player_id=player.id,
                    provider=SOURCE_PROVIDER,
                    provider_player_id=provider_id or f"source-row:{source.row_number}",
                    season=source.season,
                    season_type=SEASON_TYPE,
                    parser_version=PARSER_VERSION,
                    imported_at=imported_at,
                )
                db.add(existing)
                rows_inserted += 1
            else:
                rows_updated += 1
            try:
                _assign_row(existing, source, player, teams_by_normalized_name, source_hash, imported_at)
            except Exception as exc:  # Keep the import transactional; report before re-raising.
                errors.append({"row_number": source.row_number, "error": str(exc), "match_type": match_type})
                raise
            touched_players.add(player.id)

        run.requested_player_ids = sorted(touched_players)
        run.players_requested = len(touched_players)
        run.players_succeeded = len(touched_players)
        run.players_unmatched = len(report["unmatched_source_rows"])
        run.rows_inserted = rows_inserted
        run.rows_updated = rows_updated
        run.status = "completed"
        run.completed_at = datetime.now(timezone.utc)
        run.error_summary = {"source_sha256": source_hash, "errors": errors} if errors else {"source_sha256": source_hash}
        db.commit()
        report.update(
            {
                "run_id": run.id,
                "players_imported": len(touched_players),
                "rows_inserted": rows_inserted,
                "rows_updated": rows_updated,
                "provider_mappings_inserted": mappings_inserted,
                "provider_mappings_unchanged": mappings_unchanged,
            }
        )
        return report


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True, help="Approved CSV or XLSX export of the Season Stats tab.")
    parser.add_argument("--apply", action="store_true", help="Write verified source rows to the historical season table.")
    parser.add_argument("--sealed-manifest", type=Path, help="Complete sealed source manifest required for --apply.")
    parser.add_argument("--report", type=Path, help="Write the reconciliation report as JSON.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    path = args.input.expanduser().resolve()
    if not path.is_file():
        raise SystemExit(f"Source export does not exist: {path}")
    if args.apply:
        if args.sealed_manifest is None or not args.sealed_manifest.is_file():
            raise SystemExit("--apply requires a complete sealed source manifest.")
        manifest = json.loads(args.sealed_manifest.read_text(encoding="utf-8"))
        snapshots = manifest.get("snapshots", []) if isinstance(manifest, dict) else []
        required_workbooks = {"player_id_details", "team_rankings", "player_previous_stats", "annual_projections", "schedules", "cfb27_ratings"}
        if {item.get("workbook") for item in snapshots if isinstance(item, dict)} != required_workbooks or _file_hash(path) not in {item.get("sha256") for item in snapshots if isinstance(item, dict)}:
            raise SystemExit("--apply requires the history file to be in a complete sealed six-workbook manifest.")
    try:
        report = import_rows(path, apply=args.apply)
    except Exception as exc:
        raise SystemExit(f"Historical stats import failed without committing partial rows: {exc}") from exc
    payload = json.dumps(report, indent=2, sort_keys=True)
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(payload + "\n", encoding="utf-8")
    print(payload)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
